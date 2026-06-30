import csv
import os
import urllib.request
import urllib3
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

SRC = "movies.csv"
DST = "movies_with_posters.xlsx"
IMG_DIR = "posters"
os.makedirs(IMG_DIR, exist_ok=True)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
CELL_FONT = Font(size=11)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

with open(SRC, "r", encoding="utf-8-sig") as f:
    movies = list(csv.DictReader(f))

wb = Workbook()
ws = wb.active
ws.title = "電影列表"

headers = ["編號", "中文片名", "英文片名", "分類", "地區", "時長", "上映日期", "評分", "海報"]
col_widths = [8, 20, 30, 20, 22, 10, 14, 8, 18]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 25

for row_idx, m in enumerate(movies, 2):
    values = [
        int(m["id"]),
        m["cn_name"],
        m["en_name"],
        m["categories"],
        m["region"],
        m["duration"],
        m["release_date"],
        m["score"],
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = CELL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    img_url = m["cover"]
    try:
        req = urllib.request.Request(img_url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=10)
        img_data = BytesIO(resp.read())
        img = Image.open(img_data)
        img.thumbnail((120, 180))
        img_path = os.path.join(IMG_DIR, f"{m['id']}.png")
        img.save(img_path)

        from openpyxl.drawing.image import Image as XlImage
        xl_img = XlImage(img_path)
        xl_img.width = 90
        xl_img.height = 135

        ws.add_image(xl_img, f"I{row_idx}")
        ws.row_dimensions[row_idx].height = 140
    except Exception as e:
        ws.cell(row=row_idx, column=9, value=f"下載失敗")
        print(f"  [{m['id']}] 海報下載失敗: {e}")

ws.column_dimensions["I"].width = 18

wb.save(DST)
print(f"\n完成！已儲存至 {DST}")
